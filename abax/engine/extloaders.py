"""Loaders for closed-workbook external references — ``=[Data.xlsx]Sheet1!B2``.

:mod:`abax.core.externref` resolves ``[Book]Sheet!A1`` references by handing a
resolved :class:`~pathlib.Path` to an injected *loader* (``Loader =
Callable[[Path], Any]``) that must return an object exposing
``get_sheet(name)`` and ``.sheet``, whose sheets answer ``get_value(row, col)``.
Out of the box the hub only reaches ``.abax``/``.json`` via the engine
``Document``; this module widens the reachable formats with a suffix dispatch:

``.abax`` / ``.json``
    The existing :meth:`abax.engine.document.Document.open` path (native
    envelope or foreign exchange JSON).
``.xlsx``
    An openpyxl-backed **values-only** read. Formula cells come back as the
    cached values Excel stored in the file (openpyxl's ``data_only`` view);
    when no cached value exists (e.g. the file was written by a library that
    never calculated it) the formula *text* is returned as inert literal text —
    external formulas are never re-evaluated by abax. openpyxl is optional: if
    it is missing, :class:`ExternalLoadError` names the ``abax[excel]`` extra.
``.csv`` / ``.tsv``
    Stdlib ``csv`` into a single-sheet workbook named after the file stem, so
    both ``=[data.csv]!A1`` (empty sheet part — the first sheet) and
    ``=[data.csv]data!A1`` resolve.

Everything here is strictly read-only: source files are opened for reading
only and are never written, locked for write, or mutated.

Integration
-----------
Point :data:`abax.core.externref.ALLOWED_SUFFIXES` at
:data:`SUPPORTED_SUFFIXES` and swap ``_default_loader`` for
:func:`load_external` (keeping the lazy import so core stays pure stdlib).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from ..core.cells import Cell
from ..core.io import csv_io
from ..core.sheet import Sheet
from ..core.workbook import Workbook

#: Every file suffix :func:`load_external` can dispatch. The externref hub's
#: ``ALLOWED_SUFFIXES`` gate should point here so the two lists never drift.
SUPPORTED_SUFFIXES: tuple[str, ...] = (".abax", ".json", ".xlsx", ".csv", ".tsv")

_OPENPYXL_MSG = (
    "External .xlsx references require 'openpyxl'. Install it with:\n"
    "    pip install openpyxl\n"
    "or install abax's excel extra:  pip install abax[excel]"
)


class ExternalLoadError(Exception):
    """An external workbook could not be loaded.

    Raised for a suffix outside :data:`SUPPORTED_SUFFIXES` and for an ``.xlsx``
    reference when openpyxl is not installed (the message names the
    ``abax[excel]`` extra). I/O and parse failures from the underlying readers
    propagate as-is — the externref hub records any exception per-book and
    surfaces ``#REF!`` without crashing.
    """


def load_external(path: Path) -> Any:
    """Load the external workbook at *path*, dispatching on its suffix.

    Satisfies the :data:`abax.core.externref.Loader` contract: the returned
    workbook exposes ``get_sheet(name)`` / ``.sheet`` and its sheets answer
    ``get_value(row, col)``. The source file is only ever read.

    Raises :class:`ExternalLoadError` for an unsupported suffix, or for
    ``.xlsx`` when openpyxl is absent.
    """
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix in (".abax", ".json"):
        return _load_native(path)
    if suffix == ".xlsx":
        return _load_xlsx(path)
    if suffix in (".csv", ".tsv"):
        return _load_delimited(path, delimiter="\t" if suffix == ".tsv" else ",")
    raise ExternalLoadError(
        f"unsupported external workbook type {suffix!r} "
        f"(supported: {', '.join(SUPPORTED_SUFFIXES)})"
    )


# -- native (.abax / .json) --------------------------------------------------


def _load_native(path: Path) -> Workbook:
    """The pre-existing path: engine ``Document`` smart-load (envelope or
    foreign exchange JSON). Imported lazily so merely importing this module
    stays cheap."""
    from .document import Document

    return Document.open(str(path)).workbook


# -- delimited text (.csv / .tsv) --------------------------------------------


def _load_delimited(path: Path, delimiter: str) -> Workbook:
    """Stdlib csv reader into a single-sheet workbook.

    The sheet is named after the file stem (``data.csv`` → sheet ``data``), so
    an external ref may either name it explicitly or leave the sheet part
    empty and get the workbook's first (only) sheet.
    """
    sheet = csv_io.load_csv(path, delimiter=delimiter)  # names sheet after stem
    return Workbook.from_sheets([sheet])


# -- Excel (.xlsx), values only ----------------------------------------------


class _ValueCell(Cell):
    """A cell that holds an already-computed value and is never a formula.

    The core :class:`~abax.core.cells.Cell` stores raw *text* and re-parses it
    on read; external xlsx cells arrive as finished Python values (int, float,
    str, bool, datetime, …), so this subclass hands the stored object straight
    back from :meth:`literal`. That both preserves exact types and keeps a
    formula-text fallback string beginning with ``=`` inert — abax never
    evaluates content from a closed external workbook.
    """

    __slots__ = ()

    def __init__(self, value: Any) -> None:
        super().__init__(str(value))  # raw text kept sane for get_raw/repr
        self.value = value

    @property
    def is_formula(self) -> bool:
        return False

    def literal(self) -> Any:
        return self.value


def _is_formula_value(value: Any) -> bool:
    """Is *value* (a cell value from openpyxl's formula view) a formula?

    Plain formulas are strings beginning with ``=``; array/data-table formulas
    are openpyxl objects exposing a ``text`` attribute.
    """
    return (isinstance(value, str) and value.startswith("=")) or hasattr(value, "text")


def _formula_text(value: Any) -> str:
    """The source text of a formula value (string or openpyxl formula object)."""
    if isinstance(value, str):
        return value
    text = getattr(value, "text", None)
    return text if isinstance(text, str) else str(value)


def _load_xlsx(path: Path) -> Workbook:
    """Read *path* with openpyxl into a values-only :class:`Workbook`.

    Two read-only passes over the file: the ``data_only`` view supplies the
    cached values Excel stored for formula cells, the formula view supplies
    cell coverage plus the formula text used as a literal fallback when no
    cached value exists. Nothing is ever written back.
    """
    try:
        import openpyxl  # noqa: PLC0415 — optional dep, guarded on purpose
    except ImportError as exc:
        raise ExternalLoadError(_OPENPYXL_MSG) from exc

    wb_formulas = openpyxl.load_workbook(path, read_only=True, data_only=False)
    wb_values = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for ws_f, ws_v in zip(wb_formulas.worksheets, wb_values.worksheets):
            cached = {
                (c.row, c.column): c.value
                for row in ws_v.iter_rows()
                for c in row
                if c.value is not None
            }
            sheet = Sheet(ws_f.title)
            cells = sheet._cells  # bulk fill; _bounds_dirty is already True
            for row in ws_f.iter_rows():
                for c in row:
                    value = c.value
                    if value is None:
                        continue
                    if _is_formula_value(value):
                        cached_value = cached.get((c.row, c.column))
                        value = (cached_value if cached_value is not None
                                 else _formula_text(value))
                    cells[(c.row - 1, c.column - 1)] = _ValueCell(value)
            sheets.append(sheet)
    finally:
        # read_only workbooks hold the file handle open until closed.
        wb_formulas.close()
        wb_values.close()
    return Workbook.from_sheets(sheets)

"""Excel (.xlsx) import/export via openpyxl — optional, with a clear fallback.

If openpyxl is not installed, the loader/saver raise a descriptive
``RuntimeError`` telling the user how to enable it. This keeps the core engine
free of any hard third-party dependency (see docs/architecture.md).

Beyond raw cell text, both directions carry the fidelity model the native
envelope persists (schema v2 + formatting): per-cell number formats, visual
styles (bold/italic/underline, alignment, text/fill colours), borders, column
widths, row heights, frozen panes, merged regions, and conditional-formatting
rules. Embedded charts whose kind has a native Excel counterpart (line, bar,
scatter) become real Excel charts and come back as ChartObjects. Everything
degrades gracefully — a workbook with no styling round-trips exactly as
before, and foreign .xlsx features abax has no model for are simply left
behind rather than erroring.
"""

from __future__ import annotations

import re
from pathlib import Path

from ..core import chartobj
from ..core.format.cellstyle import CellStyle
from ..core.format.condformat import CondRule, CondStyle, parse_css
from ..core.reference import col_to_index, index_to_col, parse_a1, parse_range, to_a1
from ..core.sheet import Sheet
from ..core.workbook import Workbook

try:
    import openpyxl  # type: ignore
    from openpyxl.chart import (  # type: ignore
        BarChart,
        LineChart,
        Reference,
        ScatterChart,
        Series,
    )
    from openpyxl.formatting.rule import ColorScaleRule, Rule  # type: ignore
    from openpyxl.styles import (  # type: ignore
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.styles.differential import DifferentialStyle  # type: ignore

    HAS_OPENPYXL = True
except ImportError:  # pragma: no cover - exercised only without the dep
    openpyxl = None
    HAS_OPENPYXL = False


_FALLBACK_MSG = (
    "Excel import/export requires 'openpyxl'. Install it with:\n"
    "    pip install openpyxl\n"
    "or install abax's excel extra:  pip install abax[excel]"
)


def load_xlsx(path: str | Path) -> Workbook:
    if not HAS_OPENPYXL:
        raise RuntimeError(_FALLBACK_MSG)
    path = Path(path)
    # data_only=False keeps formulas as text so abax re-evaluates them itself.
    wb_x = openpyxl.load_workbook(path, data_only=False)
    wb = Workbook.__new__(Workbook)
    wb.sheets = []
    wb.active = 0
    for ws in wb_x.worksheets:
        sheet = Sheet(ws.title)
        # openpyxl returns formulas as "=..."; numbers/strings as-is.
        sheet.set_cells_bulk(
            (cell.row - 1, cell.column - 1, str(cell.value))
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None)
        _read_fidelity(ws, sheet)
        wb.sheets.append(sheet)
    wb._add_default_if_empty()
    # Charts last: their data references may span sheets, so every sheet must
    # already exist (and be linked) before any chart is rebuilt.
    for ws, sheet in zip(wb_x.worksheets, wb.sheets):
        _read_charts(ws, sheet, wb)
    return wb


def save_xlsx(wb: Workbook, path: str | Path, *, values: bool = False) -> None:
    """Write a workbook to .xlsx.

    ``values=False`` (default) writes raw cell text, so formulas survive the
    round-trip into Excel. ``values=True`` writes computed values instead.
    Either way the sheet's formatting fidelity (number formats, styles,
    borders, layout, merges, conditional formats) and its line/bar/scatter
    embedded charts are carried along.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError(_FALLBACK_MSG)
    path = Path(path)
    wb_x = openpyxl.Workbook()
    # Remove the default sheet openpyxl creates; we add our own.
    default = wb_x.active
    wb_x.remove(default)
    ws_by_name: dict = {}
    for sheet in wb.sheets:
        ws = wb_x.create_sheet(title=sheet.name[:31])  # Excel caps title at 31
        ws_by_name[sheet.name] = ws
        n_rows, n_cols = sheet.used_bounds()
        for r in range(n_rows):
            for c in range(n_cols):
                cell = sheet.get_cell(r, c)
                if cell is None:
                    continue
                if values:
                    val = sheet.get_value(r, c)
                    ws.cell(row=r + 1, column=c + 1, value=_coerce(val))
                else:
                    ws.cell(row=r + 1, column=c + 1, value=_excel_raw(cell))
        _write_fidelity(ws, sheet)
    # Charts last: a chart may source a sheet created after its own.
    for sheet in wb.sheets:
        if sheet.charts:
            _write_charts(ws_by_name[sheet.name], sheet, wb, ws_by_name)
    if not wb_x.worksheets:
        wb_x.create_sheet(title="Sheet1")
    wb_x.save(path)


def _excel_raw(cell) -> object:
    """Convert a abax raw cell to a value openpyxl will write faithfully."""
    if cell.is_formula:
        return cell.raw  # already begins with '='
    return cell.literal()


def _coerce(val: object) -> object:
    from ..core.errors import CellError

    if isinstance(val, CellError):
        return str(val)
    return val


# --- formatting fidelity ----------------------------------------------------
#
# Everything below maps the sheet's fidelity model (the same fields the native
# envelope persists — see Workbook.to_envelope) onto openpyxl's, in both
# directions. Each feature is independent and sparse: an unstyled workbook
# writes none of it, and unmappable foreign styling is skipped, never fatal.

# abax number-format specs (core.format.cellformat) <-> Excel format codes.
_SPEC_TO_NUMFMT = {
    "general": "General",
    "int": "0",
    "comma": "#,##0.00",
    "currency": "$#,##0.00",
    "percent": "0.00%",
    "sci": "0.000E+00",  # abax's sci displays 3 decimals (f"{v:.3e}")
    "text": "@",
}

# Same shape cellformat's _FIXED accepts: "fixed2" or "fixed:2".
_FIXED_SPEC = re.compile(r"^fixed:?(\d+)$")
_FIXED_CODE = re.compile(r"^0\.(0+)$")


def _spec_to_numfmt(spec: str) -> "str | None":
    """abax format spec -> Excel number-format code (None = leave General)."""
    spec = (spec or "").lower()
    if spec in _SPEC_TO_NUMFMT:
        return _SPEC_TO_NUMFMT[spec]
    m = _FIXED_SPEC.match(spec)
    if m:
        n = int(m.group(1))
        return "0." + "0" * n if n else "0"
    return None


def _numfmt_to_spec(code: str) -> "str | None":
    """Excel number-format code -> abax spec, best effort.

    Exact inverses of :data:`_SPEC_TO_NUMFMT` come first, then heuristics for
    foreign codes (a ``$`` means currency, a ``%`` percent, …). Codes with no
    abax counterpart (dates, custom masks) return None and display as general.
    """
    if not code or code.lower() == "general":
        return None
    if code == "@":
        return "text"
    if "$" in code or "[$" in code:
        return "currency"
    if "%" in code:
        return "percent"
    up = code.upper()
    if "E+" in up or "E-" in up:
        return "sci"
    if "#,##" in code:
        return "comma"
    m = _FIXED_CODE.match(code)
    if m:
        return f"fixed{len(m.group(1))}"
    if code == "0":
        return "int"
    return None


def _argb(hex_color: str) -> str:
    """``'#rrggbb'`` -> ``'FFRRGGBB'`` (the opaque-ARGB form openpyxl uses)."""
    return "FF" + hex_color.lstrip("#").upper()


def _hex_of(color) -> str:
    """openpyxl Color -> ``'#rrggbb'``, or ``''`` when it isn't plain RGB.

    Theme/indexed colours and fully transparent ARGB (alpha ``00``) come back
    empty — abax's model only holds concrete RGB values.
    """
    if color is None or getattr(color, "type", None) != "rgb":
        return ""
    rgb = color.rgb
    if not isinstance(rgb, str) or len(rgb) not in (6, 8):
        return ""
    if len(rgb) == 8 and rgb[:2] == "00":
        return ""
    return "#" + rgb[-6:].lower()


# --- cell styles (CellStyle <-> Font/Alignment/PatternFill) -----------------


def _apply_style(cell, st: CellStyle) -> None:
    """Stamp a :class:`CellStyle` onto an openpyxl cell (empty style = no-op)."""
    if st.bold or st.italic or st.underline or st.text_color:
        cell.font = Font(
            bold=st.bold or None,
            italic=st.italic or None,
            underline="single" if st.underline else None,
            color=_argb(st.text_color) if st.text_color else None)
    if st.align:
        cell.alignment = Alignment(horizontal=st.align)
    if st.bg_color:
        cell.fill = PatternFill(fill_type="solid", start_color=_argb(st.bg_color))


def _read_style(cell) -> "CellStyle | None":
    """The :class:`CellStyle` equivalent of an openpyxl cell's styling, or None."""
    kw: dict = {}
    font = cell.font
    if font.bold:
        kw["bold"] = True
    if font.italic:
        kw["italic"] = True
    if font.underline and font.underline != "none":
        kw["underline"] = True
    text = _hex_of(font.color)
    if text:
        kw["text_color"] = text
    horiz = cell.alignment.horizontal
    if horiz in ("left", "center", "right"):
        kw["align"] = horiz
    fill = cell.fill
    if fill.patternType == "solid":
        bg = _hex_of(fill.start_color)
        if bg:
            kw["bg_color"] = bg
    return CellStyle(**kw) if kw else None


# --- borders -----------------------------------------------------------------

_EDGES = ("top", "bottom", "left", "right")

# abax's three border weights are valid openpyxl side styles verbatim; foreign
# styles fold to the nearest weight on import.
_ABAX_SIDES = ("thin", "medium", "thick")
_SIDE_TO_ABAX = {
    "hair": "thin", "thin": "thin", "dotted": "thin", "dashed": "thin",
    "dashDot": "thin", "dashDotDot": "thin",
    "medium": "medium", "mediumDashed": "medium", "mediumDashDot": "medium",
    "mediumDashDotDot": "medium", "slantDashDot": "medium", "double": "medium",
    "thick": "thick",
}


def _write_border(edges: dict) -> "Border | None":
    sides = {e: Side(style=s) for e, s in edges.items()
             if e in _EDGES and s in _ABAX_SIDES}
    return Border(**sides) if sides else None


def _read_border(border) -> dict:
    edges = {}
    for e in _EDGES:
        style = getattr(getattr(border, e, None), "style", None)
        if style is not None:
            edges[e] = _SIDE_TO_ABAX.get(style, "thin")
    return edges


# --- layout units ------------------------------------------------------------
#
# abax stores column widths and row heights in *pixels* (Qt view geometry).
# Excel's units differ per axis:
#   - column width is in characters of the default font's max digit width
#     (Calibri 11 ≈ 7 px at 96 DPI) with ~5 px of cell padding, so
#     chars = (px - 5) / 7 and px = chars * 7 + 5;
#   - row height is in points, so pt = px * 72/96 = px * 0.75.
# Both conversions invert exactly after round(), so a round-trip returns the
# original pixel value.

_COL_PAD_PX = 5
_COL_CHAR_PX = 7.0
_ROW_PT_PER_PX = 0.75


def _col_px_to_chars(px: int) -> float:
    return max((px - _COL_PAD_PX) / _COL_CHAR_PX, 0.0)


def _col_chars_to_px(chars: float) -> int:
    return round(chars * _COL_CHAR_PX + _COL_PAD_PX)


def _row_px_to_pt(px: int) -> float:
    return px * _ROW_PT_PER_PX


def _row_pt_to_px(pt: float) -> int:
    return round(pt / _ROW_PT_PER_PX)


# --- conditional formatting ---------------------------------------------------

_CELLIS_OPS = {
    ">": "greaterThan", "<": "lessThan",
    ">=": "greaterThanOrEqual", "<=": "lessThanOrEqual",
    "==": "equal", "!=": "notEqual", "between": "between",
}
_OPS_CELLIS = {v: k for k, v in _CELLIS_OPS.items()}

_TEXT_KINDS = {
    "contains": "containsText",
    "beginswith": "beginsWith",
    "endswith": "endsWith",
}
_KINDS_TEXT = {v: k for k, v in _TEXT_KINDS.items()}


def _cf_operand(v) -> str:
    """A CondRule threshold as a cellIs formula operand (text gets quoted)."""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return repr(v)
    s = str(v)
    try:
        float(s)
        return s  # numeric-looking text passes through bare
    except ValueError:
        return '"' + s.replace('"', '""') + '"'


def _cf_parse_operand(s):
    """Inverse of :func:`_cf_operand`: unquote text, re-type numbers."""
    if s is None:
        return None
    s = str(s)
    if len(s) >= 2 and s.startswith('"') and s.endswith('"'):
        return s[1:-1].replace('""', '"')
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return s


def _write_dxf(rule: CondRule) -> "DifferentialStyle | None":
    """The differential style a matched rule applies (mirrors condformat)."""
    if rule.css:
        st = parse_css(rule.css)
    else:
        st = CondStyle(fill=(rule.color or "").lower() or None)
    fill = font = None
    if st.fill:
        fill = PatternFill(fill_type="solid", start_color=_argb(st.fill))
    if st.text or st.bold or st.italic or st.underline:
        font = Font(
            bold=st.bold or None,
            italic=st.italic or None,
            underline="single" if st.underline else None,
            color=_argb(st.text) if st.text else None)
    if fill is None and font is None:
        return None
    return DifferentialStyle(fill=fill, font=font)


def _read_dxf(dxf) -> "tuple[str, str]":
    """A loaded differential style as CondRule fields: ``(color, css)``.

    A plain solid fill maps to the rule's ``color``; anything richer (text
    colour / emphasis) becomes a ``css`` declaration, the same vocabulary
    ``condformat.parse_css`` understands.
    """
    if dxf is None:
        return "", ""
    fill = ""
    if dxf.fill is not None:
        fill = _hex_of(dxf.fill.fgColor) or _hex_of(dxf.fill.bgColor)
    text = ""
    bold = italic = underline = False
    font = dxf.font
    if font is not None:
        bold = bool(font.b)
        italic = bool(font.i)
        underline = bool(font.u) and font.u != "none"
        text = _hex_of(font.color)
    if not (text or bold or italic or underline):
        return fill, ""
    parts = []
    if fill:
        parts.append(f"background: {fill}")
    if text:
        parts.append(f"color: {text}")
    if bold:
        parts.append("font-weight: bold")
    if italic:
        parts.append("font-style: italic")
    if underline:
        parts.append("text-decoration: underline")
    return "", "; ".join(parts)


def _write_cond_rule(rule: CondRule):
    """One CondRule -> ``(sqref, openpyxl Rule | None)``.

    None means the kind has no .xlsx counterpart (``regex``) — the rule is
    skipped, never fatal. Colour-scale rules carry their gradient in the rule
    itself; every other kind carries a differential style.
    """
    r1, c1, r2, c2 = parse_range(rule.range)
    sqref = f"{to_a1(r1, c1)}:{to_a1(r2, c2)}"
    anchor = to_a1(r1, c1)  # relative top-left, the cell text formulas test
    kind = rule.kind
    dxf = _write_dxf(rule)
    if kind in _CELLIS_OPS:
        formula = [_cf_operand(rule.value)]
        if kind == "between":
            formula.append(_cf_operand(rule.value2))
        return sqref, Rule(type="cellIs", operator=_CELLIS_OPS[kind],
                           formula=formula, dxf=dxf)
    if kind == "colorscale":
        return sqref, ColorScaleRule(
            start_type="min", start_color=_argb(str(rule.value)),
            end_type="max", end_color=_argb(str(rule.value2)))
    if kind == "colorscale3":
        return sqref, ColorScaleRule(
            start_type="min", start_color=_argb(str(rule.value)),
            mid_type="percentile", mid_value=50, mid_color=_argb(str(rule.color)),
            end_type="max", end_color=_argb(str(rule.value2)))
    if kind in _TEXT_KINDS:
        text = str(rule.value)
        esc = text.replace('"', '""')
        if kind == "contains":
            f = f'NOT(ISERROR(SEARCH("{esc}",{anchor})))'
        elif kind == "beginswith":
            f = f'LEFT({anchor},{len(text)})="{esc}"'
        else:
            f = f'RIGHT({anchor},{len(text)})="{esc}"'
        t = _TEXT_KINDS[kind]
        return sqref, Rule(type=t, operator=t, text=text, formula=[f], dxf=dxf)
    if kind == "blank":
        return sqref, Rule(type="containsBlanks",
                           formula=[f"LEN(TRIM({anchor}))=0"], dxf=dxf)
    if kind == "notblank":
        return sqref, Rule(type="notContainsBlanks",
                           formula=[f"LEN(TRIM({anchor}))>0"], dxf=dxf)
    if kind in ("above_avg", "below_avg"):
        return sqref, Rule(type="aboveAverage", dxf=dxf,
                           aboveAverage=None if kind == "above_avg" else False)
    if kind in ("top_n", "bottom_n", "top_pct", "bottom_pct"):
        try:
            rank = int(float(str(rule.value)))  # Excel ranks are integers
        except (TypeError, ValueError):
            return sqref, None
        return sqref, Rule(type="top10", rank=rank, dxf=dxf,
                           percent=True if kind.endswith("_pct") else None,
                           bottom=True if kind.startswith("bottom") else None)
    if kind == "duplicate":
        return sqref, Rule(type="duplicateValues", dxf=dxf)
    if kind == "unique":
        return sqref, Rule(type="uniqueValues", dxf=dxf)
    return sqref, None  # regex & future kinds: no .xlsx counterpart


def _read_cond_rule(rng: str, xr) -> "CondRule | None":
    """One openpyxl rule over one range -> CondRule (None = unmappable type)."""
    t = xr.type
    color, css = _read_dxf(xr.dxf)
    kw: dict = {"css": css} if css else {}
    if color:
        kw["color"] = color
    if t == "cellIs" and xr.operator in _OPS_CELLIS:
        vals = [_cf_parse_operand(f) for f in xr.formula]
        return CondRule(range=rng, kind=_OPS_CELLIS[xr.operator],
                        value=vals[0] if vals else None,
                        value2=vals[1] if len(vals) > 1 else None, **kw)
    if t == "colorScale" and xr.colorScale is not None:
        colors = [_hex_of(c) for c in xr.colorScale.color]
        if len(colors) == 2 and all(colors):
            return CondRule(range=rng, kind="colorscale",
                            value=colors[0], value2=colors[1])
        if len(colors) == 3 and all(colors):
            return CondRule(range=rng, kind="colorscale3",
                            value=colors[0], value2=colors[2], color=colors[1])
        return None
    if t in _KINDS_TEXT:
        return CondRule(range=rng, kind=_KINDS_TEXT[t], value=xr.text, **kw)
    if t == "containsBlanks":
        return CondRule(range=rng, kind="blank", **kw)
    if t == "notContainsBlanks":
        return CondRule(range=rng, kind="notblank", **kw)
    if t == "aboveAverage":
        kind = "above_avg" if xr.aboveAverage in (None, True) else "below_avg"
        return CondRule(range=rng, kind=kind, **kw)
    if t == "top10":
        kind = (("bottom" if xr.bottom else "top")
                + ("_pct" if xr.percent else "_n"))
        return CondRule(range=rng, kind=kind, value=int(xr.rank or 10), **kw)
    if t == "duplicateValues":
        return CondRule(range=rng, kind="duplicate", **kw)
    if t == "uniqueValues":
        return CondRule(range=rng, kind="unique", **kw)
    return None  # dataBar, iconSet, expression, … — no abax counterpart


# --- whole-sheet fidelity, both directions ------------------------------------


def _write_fidelity(ws, sheet: Sheet) -> None:
    """Copy a sheet's formatting fidelity onto an openpyxl worksheet.

    Runs after the cell values are written; every feature is sparse, so an
    unstyled sheet writes nothing. Merges come last so per-cell styling always
    targets regular cells, never MergedCell proxies.
    """
    for (r, c), spec in sheet.cell_formats.items():
        code = _spec_to_numfmt(spec)
        if code is not None:
            ws.cell(row=r + 1, column=c + 1).number_format = code
    for (r, c), st in sheet.cell_styles.items():
        _apply_style(ws.cell(row=r + 1, column=c + 1), st)
    for (r, c), edges in sheet.cell_borders.items():
        border = _write_border(edges)
        if border is not None:
            ws.cell(row=r + 1, column=c + 1).border = border
    for c, px in sheet.col_widths.items():
        if px > 0:
            ws.column_dimensions[index_to_col(c)].width = _col_px_to_chars(px)
    for r, px in sheet.row_heights.items():
        if px > 0:
            ws.row_dimensions[r + 1].height = _row_px_to_pt(px)
    for (r1, c1, r2, c2) in sheet.merges:
        ws.merge_cells(start_row=r1 + 1, start_column=c1 + 1,
                       end_row=r2 + 1, end_column=c2 + 1)
    if sheet.frozen_rows or sheet.frozen_cols:
        # freeze_panes names the first *unfrozen* cell, e.g. 1 row + 0 cols = A2.
        ws.freeze_panes = to_a1(sheet.frozen_rows, sheet.frozen_cols)
    for rule in sheet.cond_rules:
        try:
            sqref, xrule = _write_cond_rule(rule)
        except Exception:
            continue  # a malformed rule degrades to "not exported", never fatal
        if xrule is not None:
            ws.conditional_formatting.add(sqref, xrule)


def _read_fidelity(ws, sheet: Sheet) -> None:
    """Copy an openpyxl worksheet's formatting into the sheet's fidelity model."""
    for row in ws.iter_rows():
        for cell in row:
            if not cell.has_style:
                continue
            key = (cell.row - 1, cell.column - 1)
            spec = _numfmt_to_spec(cell.number_format)
            if spec is not None:
                sheet.cell_formats[key] = spec
            st = _read_style(cell)
            if st is not None:
                sheet.cell_styles[key] = st
            edges = _read_border(cell.border)
            if edges:
                sheet.cell_borders[key] = edges
    for key, dim in ws.column_dimensions.items():
        if not dim.width or dim.width <= 0:
            continue
        px = _col_chars_to_px(dim.width)
        # A <col> entry can span several columns (min..max, 1-based).
        lo, hi = dim.min, dim.max
        if lo is None or hi is None:
            lo = hi = col_to_index(key) + 1
        for cx in range(lo, hi + 1):
            sheet.col_widths[cx - 1] = px
    for idx, dim in ws.row_dimensions.items():
        if dim.height and dim.height > 0:
            sheet.row_heights[int(idx) - 1] = _row_pt_to_px(dim.height)
    fp = ws.freeze_panes
    if fp and str(fp) != "A1":
        r, c = parse_a1(str(fp))
        sheet.frozen_rows, sheet.frozen_cols = r, c
    sheet.merges = [
        (rng.min_row - 1, rng.min_col - 1, rng.max_row - 1, rng.max_col - 1)
        for rng in ws.merged_cells.ranges]
    rules = []
    for cf in ws.conditional_formatting:
        for rng in str(cf.sqref).split():
            for xr in cf.rules:
                try:
                    rule = _read_cond_rule(rng, xr)
                except Exception:
                    continue  # foreign rule shapes degrade to "not imported"
                if rule is not None:
                    rules.append(rule)
    sheet.cond_rules = rules


# --- embedded charts -----------------------------------------------------------
#
# ChartObjects whose kind has a native Excel counterpart (line, bar, scatter)
# become real openpyxl charts, and native charts of those types come back as
# ChartObjects. Both directions follow chart_data's column pairing
# (core/chartobj.py) so Excel draws the same series abax renders. Kinds Excel
# can't express through openpyxl (histogram, box, violin, qq, ecdf, heatmap,
# waterfall) are skipped cleanly on export — the cell data still lands, the
# picture survives only in the native envelope — and foreign chart types abax
# has no model for (pie, area, 3-D, …) are ignored on import.

# openpyxl chart sizes are centimetres; anchors carry EMU. abax stores pixels
# (96 DPI): px * 2.54/96 cm, and 1 px = 9525 EMU (914400 EMU/inch ÷ 96), so a
# px → cm → EMU → px trip inverts exactly after round().
_CM_PER_PX = 2.54 / 96.0
_EMU_PER_PX = 9525


def _px_to_cm(px: int) -> float:
    return px * _CM_PER_PX


def _emu_to_px(emu: int) -> int:
    return round(emu / _EMU_PER_PX)


def _title_text(title) -> str:
    """The plain text of an openpyxl chart Title (may span rich-text runs)."""
    rich = getattr(getattr(title, "tx", None), "rich", None)
    if rich is None:
        return ""
    return "".join(r.t for p in (rich.p or [])
                   for r in (getattr(p, "r", None) or []) if r.t)


def _ref_box(src) -> "tuple[str, int, int, int, int] | None":
    """A series data source -> its ``(sheet, r1, c1, r2, c2)`` box, or None.

    ``src`` is any openpyxl holder with a ``numRef``/``strRef`` formula —
    series values, categories, scatter x-values, or a from-data series title.
    Coordinates come back zero-based and normalized; a literal (non-reference)
    source returns None.
    """
    if src is None:
        return None
    ref = getattr(src, "numRef", None) or getattr(src, "strRef", None)
    body = (getattr(ref, "f", "") or "").strip()
    if not body:
        return None
    sheet = ""
    if "!" in body:
        sheet, body = body.rsplit("!", 1)
        sheet = sheet.strip().strip("'").replace("''", "'")
    body = body.replace("$", "")
    a, _, b = body.partition(":")
    try:
        r1, c1 = parse_a1(a)
        r2, c2 = parse_a1(b) if b else (r1, c1)
    except Exception:
        return None
    return (sheet, min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))


def _range_str(sheet_name: str, r1: int, c1: int, r2: int, c2: int,
               host_name: str) -> str:
    """A box as a ChartObject range: bare on the host sheet, qualified off it."""
    body = f"{to_a1(r1, c1)}:{to_a1(r2, c2)}"
    if sheet_name and sheet_name != host_name:
        return f"{sheet_name}!{body}"
    return body


def _anchor_cell(anchor) -> "tuple[int, int]":
    """An openpyxl drawing anchor -> abax's zero-based ``(row, col)``."""
    frm = getattr(anchor, "_from", None)
    if frm is not None:
        return int(frm.row), int(frm.col)
    try:
        return parse_a1(str(anchor))  # unsaved charts anchor as "E5" strings
    except Exception:
        return (0, 0)


def _anchor_size(anchor) -> "tuple[int, int]":
    """An anchor's extent in pixels; abax's defaults when it names none.

    One-cell anchors (what abax writes, and what ``ws.add_chart(c, "E5")``
    produces) carry an EMU extent; two-cell anchors size by cell span instead,
    so those fall back to ChartObject's default 480x320.
    """
    ext = getattr(anchor, "ext", None)
    cx = getattr(ext, "cx", 0) or 0
    cy = getattr(ext, "cy", 0) or 0
    if cx and cy:
        return _emu_to_px(cx), _emu_to_px(cy)
    return 480, 320


def _write_chart(chart, wb: Workbook, host: Sheet, ws_by_name: dict):
    """One ChartObject -> an openpyxl chart ready to anchor (None = no map).

    The Reference layout mirrors ``chart_data``'s pairing: a line chart's
    columns are its series (the first column becoming the x/category axis
    under ``first_col_x``), a bar chart with a leading text column uses it as
    the category axis with the next column as values, and a scatter pairs the
    first column (x) with the second (y). When the source's first row is a
    header it feeds the series titles rather than the data.
    """
    if chart.kind not in ("line", "bar", "scatter"):
        return None
    src_sheet, a, b = chartobj._split_range(chart.source, host.name)
    ws_src = ws_by_name.get(src_sheet)
    if ws_src is None:
        return None
    grid = chartobj._load_grid(wb, host.name, chart.source)  # ChartError -> skip
    header = chartobj._has_header(grid)
    r1, c1 = parse_a1(a)
    r2, c2 = parse_a1(b)
    if r2 < r1:
        r1, r2 = r2, r1
    if c2 < c1:
        c1, c2 = c2, c1
    body_r1 = r1 + 1 if header else r1

    def ref(cc1: int, cc2: int, rr1: int = r1) -> "Reference":
        return Reference(ws_src, min_col=cc1 + 1, min_row=rr1 + 1,
                         max_col=cc2 + 1, max_row=r2 + 1)

    if chart.kind == "scatter":
        if c2 == c1:
            return None  # chart_data pairs column 1 (x) with column 2 (y)
        x = ScatterChart()
        x.scatterStyle = "marker"  # abax scatters are point clouds
        x.series.append(Series(ref(c1 + 1, c1 + 1), xvalues=ref(c1, c1, body_r1),
                               title_from_data=header))
    elif chart.kind == "line":
        x = LineChart()
        if chart.options.get("first_col_x") and c2 > c1:
            x.add_data(ref(c1 + 1, c2), titles_from_data=header)
            x.set_categories(ref(c1, c1, body_r1))
        else:
            x.add_data(ref(c1, c2), titles_from_data=header)
    else:  # bar
        x = BarChart()
        body = grid[1:] if header else grid
        first_col_text = (body and c2 > c1 and all(
            chartobj._num(row[0]) is None and row[0] not in (None, "")
            for row in body if row))
        if first_col_text:
            x.add_data(ref(c1 + 1, c1 + 1), titles_from_data=header)
            x.set_categories(ref(c1, c1, body_r1))
        else:
            x.add_data(ref(c1, c2), titles_from_data=header)
            if chart.labels:  # an explicit labels range is the category axis
                ls, la, lb = chartobj._split_range(chart.labels, host.name)
                ws_lab = ws_by_name.get(ls)
                if ws_lab is not None:
                    lr1, lc1 = parse_a1(la)
                    lr2, lc2 = parse_a1(lb)
                    x.set_categories(Reference(
                        ws_lab,
                        min_col=min(lc1, lc2) + 1, min_row=min(lr1, lr2) + 1,
                        max_col=max(lc1, lc2) + 1, max_row=max(lr1, lr2) + 1))
    if chart.title:
        x.title = chart.title
    x.width = _px_to_cm(chart.width)
    x.height = _px_to_cm(chart.height)
    return x


def _read_chart(xchart, host: Sheet, wb: Workbook):
    """One openpyxl chart -> ChartObject (None = no abax counterpart).

    The source range is rebuilt as the bounding box of the series' value
    references (plus scatter x-values and from-data title cells, so a header
    row rejoins its data). A category reference hugging the box's left edge is
    folded in — that's where chart-with-label-column layouts keep it — turning
    back into ``first_col_x`` for a line chart when the column is numeric; a
    detached one becomes the chart's ``labels`` range.
    """
    if isinstance(xchart, LineChart):
        kind = "line"
    elif isinstance(xchart, BarChart):
        kind = "bar"
    elif isinstance(xchart, ScatterChart):
        kind = "scatter"
    else:
        return None
    boxes, cats = [], []
    for s in xchart.series:
        for src in (s.val, s.yVal, s.xVal, s.tx):
            box = _ref_box(src)
            if box is not None:
                boxes.append(box)
        box = _ref_box(s.cat)
        if box is not None:
            cats.append(box)
    if not boxes:
        return None
    src_sheet = boxes[0][0]
    boxes = [bx for bx in boxes if bx[0] == src_sheet]
    r1 = min(bx[1] for bx in boxes)
    c1 = min(bx[2] for bx in boxes)
    r2 = max(bx[3] for bx in boxes)
    c2 = max(bx[4] for bx in boxes)
    labels = ""
    options: dict = {}
    cb = cats[0] if cats else None
    if cb is not None:
        if cb[0] == src_sheet and cb[4] == c1 - 1 and cb[1] >= r1 and cb[3] <= r2:
            c1 = cb[2]
            if kind == "line" and _looks_numeric(wb, src_sheet, cb[1], cb[2]):
                options["first_col_x"] = True
        else:
            labels = _range_str(cb[0], cb[1], cb[2], cb[3], cb[4], host.name)
    width, height = _anchor_size(xchart.anchor)
    return chartobj.ChartObject(
        id="", kind=kind,
        source=_range_str(src_sheet, r1, c1, r2, c2, host.name),
        title=_title_text(xchart.title), labels=labels,
        anchor=_anchor_cell(xchart.anchor), width=width, height=height,
        options=options)


def _looks_numeric(wb: Workbook, sheet_name: str, r: int, c: int) -> bool:
    """Does the first body cell of a folded-in category column hold a number?

    Numeric x categories mean the chart plotted true x/y pairs (abax's
    ``first_col_x``); text categories are plain 1..n lines with axis labels.
    """
    sheet = wb.get_sheet(sheet_name)
    return sheet is not None and chartobj._num(sheet.get_raw(r, c)) is not None


def _write_charts(ws, sheet: Sheet, wb: Workbook, ws_by_name: dict) -> None:
    """Emit the sheet's embedded charts that map to native Excel chart types.

    Mirrors the cond-format policy: anything unmappable — a kind with no Excel
    counterpart, a dead range, a missing source sheet — degrades to "not
    exported", never fatal; the cell data always lands.
    """
    for chart in sheet.charts:
        try:
            xchart = _write_chart(chart, wb, sheet, ws_by_name)
        except Exception:
            continue
        if xchart is not None:
            ws.add_chart(xchart, to_a1(*chart.anchor))


def _read_charts(ws, sheet: Sheet, wb: Workbook) -> None:
    """Rebuild ChartObjects from a worksheet's native charts (mapped types)."""
    for xchart in getattr(ws, "_charts", ()):
        try:
            chart = _read_chart(xchart, sheet, wb)
        except Exception:
            continue  # foreign chart shapes degrade to "not imported"
        if chart is not None:
            chart.id = chartobj.new_chart_id(sheet.charts)
            sheet.charts.append(chart)

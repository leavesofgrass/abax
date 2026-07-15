"""Document façade: open/save dispatch by extension; Excel when available.

Also home to the .xlsx formatting-fidelity round-trips (excel_io): number
formats, styles, borders, layout, merges, conditional formats, and embedded
charts each get a save → load → compare test. Those all
`importorskip("openpyxl")` so the thin CI environment (no openpyxl) skips
them cleanly.
"""

from __future__ import annotations

import pytest

from abax.core.workbook import Workbook
from abax.engine import HAS_OPENPYXL
from abax.engine.document import Document


def test_open_csv_and_save_json(tmp_path):
    src = tmp_path / "in.csv"
    src.write_text("1,2,=A1+B1\n")
    doc = Document.open(src)
    assert doc.workbook.sheet.get("C1") == 3
    out = tmp_path / "out.abax"
    doc.save(out)
    assert out.exists()
    reopened = Document.open(out)
    assert reopened.workbook.sheet.get("C1") == 3


def test_unsupported_extension(tmp_path):
    p = tmp_path / "x.foo"
    p.write_text("nope")
    with pytest.raises(ValueError):
        Document.open(p)


@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
def test_xlsx_roundtrip(tmp_path):
    src = tmp_path / "in.csv"
    src.write_text("5,10,=A1+B1\n")
    doc = Document.open(src)
    xlsx = tmp_path / "book.xlsx"
    doc.save(xlsx)
    assert xlsx.exists()
    reopened = Document.open(xlsx)
    # Formula survives the round-trip and re-evaluates.
    assert reopened.workbook.sheet.get("C1") == 15


@pytest.mark.skipif(HAS_OPENPYXL, reason="openpyxl IS installed")
def test_xlsx_without_openpyxl_raises(tmp_path):
    from abax.engine.excel_io import load_xlsx

    with pytest.raises(RuntimeError):
        load_xlsx(tmp_path / "missing.xlsx")


# --- .xlsx formatting fidelity (save -> load -> compare, per feature) --------


def _xlsx_roundtrip(wb, tmp_path):
    from abax.engine.excel_io import load_xlsx, save_xlsx

    p = tmp_path / "fidelity.xlsx"
    save_xlsx(wb, p)
    return load_xlsx(p).sheet


def test_xlsx_number_formats_roundtrip(tmp_path):
    pytest.importorskip("openpyxl")
    wb = Workbook()
    s = wb.sheet
    s.set("A1", "1234.5")
    for key, spec in {(0, 0): "comma", (1, 0): "fixed3", (2, 0): "int",
                      (3, 0): "currency", (4, 0): "percent", (5, 0): "sci",
                      (6, 0): "text"}.items():
        s.cell_formats[key] = spec
    s.cell_formats[(0, 5)] = "percent"  # on an empty cell beyond the data
    s2 = _xlsx_roundtrip(wb, tmp_path)
    assert s2.cell_formats == s.cell_formats


def test_xlsx_foreign_number_formats_map_best_effort(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    from abax.engine.excel_io import load_xlsx

    wb_x = openpyxl.Workbook()
    ws = wb_x.active
    for row, (value, code) in enumerate([
            (1.5, "0.00"),                        # -> fixed2
            (2.5, "$#,##0.00_);($#,##0.00)"),      # -> currency
            (3.5, "#,##0"),                        # -> comma
            (4.5, "yyyy-mm-dd"),                   # no counterpart -> dropped
    ], start=1):
        c = ws.cell(row=row, column=1, value=value)
        c.number_format = code
    p = tmp_path / "foreign.xlsx"
    wb_x.save(p)
    s = load_xlsx(p).sheet
    assert s.cell_formats[(0, 0)] == "fixed2"
    assert s.cell_formats[(1, 0)] == "currency"
    assert s.cell_formats[(2, 0)] == "comma"
    assert (3, 0) not in s.cell_formats  # dates degrade to general, not junk


def test_xlsx_cell_styles_roundtrip(tmp_path):
    pytest.importorskip("openpyxl")
    from abax.core.format.cellstyle import CellStyle

    wb = Workbook()
    s = wb.sheet
    s.set("A1", "x")
    s.cell_styles[(0, 0)] = CellStyle(bold=True, italic=True, underline=True,
                                      align="center", text_color="#112233",
                                      bg_color="#ffee00")
    s.cell_styles[(1, 1)] = CellStyle(align="right")   # single-field style
    s.cell_styles[(2, 0)] = CellStyle(bg_color="#a6e3a1")
    s2 = _xlsx_roundtrip(wb, tmp_path)
    assert s2.cell_styles == s.cell_styles


def test_xlsx_borders_roundtrip(tmp_path):
    pytest.importorskip("openpyxl")
    wb = Workbook()
    s = wb.sheet
    s.set("A1", "x")
    s.set_cell_border(0, 0, {"top": "thin", "bottom": "thick", "left": "medium"})
    s.set_cell_border(3, 2, {"right": "thin"})  # border on an empty cell
    s2 = _xlsx_roundtrip(wb, tmp_path)
    assert s2.cell_borders == s.cell_borders


def test_xlsx_foreign_border_styles_fold_to_weights(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.styles import Border, Side

    from abax.engine.excel_io import load_xlsx

    wb_x = openpyxl.Workbook()
    ws = wb_x.active
    ws.cell(row=1, column=1, value="x").border = Border(
        top=Side(style="hair"), bottom=Side(style="double"),
        left=Side(style="dashed"))
    p = tmp_path / "foreign.xlsx"
    wb_x.save(p)
    s = load_xlsx(p).sheet
    assert s.cell_border(0, 0) == {"top": "thin", "bottom": "medium", "left": "thin"}


def test_xlsx_col_widths_row_heights_roundtrip(tmp_path):
    pytest.importorskip("openpyxl")
    wb = Workbook()
    s = wb.sheet
    s.set("A1", "x")
    s.set_col_width(0, 140)
    s.set_col_width(3, 61)
    s.set_row_height(1, 30)
    s.set_row_height(5, 44)
    s2 = _xlsx_roundtrip(wb, tmp_path)
    # px -> Excel units -> px inverts exactly after round()
    assert s2.col_widths == {0: 140, 3: 61}
    assert s2.row_heights == {1: 30, 5: 44}


def test_xlsx_frozen_panes_roundtrip(tmp_path):
    pytest.importorskip("openpyxl")
    wb = Workbook()
    s = wb.sheet
    s.set("A1", "x")
    s.set_frozen(1, 2)
    s2 = _xlsx_roundtrip(wb, tmp_path)
    assert (s2.frozen_rows, s2.frozen_cols) == (1, 2)

    wb.sheet.set_frozen(3, 0)  # rows only
    s3 = _xlsx_roundtrip(wb, tmp_path)
    assert (s3.frozen_rows, s3.frozen_cols) == (3, 0)


def test_xlsx_merges_roundtrip(tmp_path):
    pytest.importorskip("openpyxl")
    wb = Workbook()
    s = wb.sheet
    s.set("A1", "anchor")
    s.merge_cells(0, 0, 1, 1)
    s.merge_cells(3, 2, 3, 4)
    s2 = _xlsx_roundtrip(wb, tmp_path)
    assert sorted(s2.merges) == [(0, 0, 1, 1), (3, 2, 3, 4)]
    assert s2.get_raw(0, 0) == "anchor"


def test_xlsx_condformat_roundtrip(tmp_path):
    pytest.importorskip("openpyxl")
    from abax.core.format.condformat import CondRule

    wb = Workbook()
    s = wb.sheet
    s.set("A1", "7")
    s.cond_rules = [
        CondRule(range="A1:A9", kind=">", value=5, color="#cc0000"),
        CondRule(range="B1:B9", kind="between", value=2, value2=5.5),
        CondRule(range="C1:C9", kind="colorscale", value="#0000ff", value2="#00ff00"),
        CondRule(range="D1:D9", kind="colorscale3", value="#0000ff",
                 value2="#00ff00", color="#ffffff"),
        CondRule(range="E1:E9", kind="contains", value="foo"),
        CondRule(range="F1:F9", kind="blank"),
        CondRule(range="G1:G9", kind="above_avg"),
        CondRule(range="H1:H9", kind="top_n", value=3),
        CondRule(range="I1:I9", kind="bottom_pct", value=10),
        CondRule(range="J1:J9", kind="unique"),
        CondRule(range="K1:K9", kind="==", value="foo"),
    ]
    s2 = _xlsx_roundtrip(wb, tmp_path)
    by_range = {r.range: r for r in s2.cond_rules}
    assert by_range["A1:A9"].kind == ">" and by_range["A1:A9"].value == 5
    assert by_range["A1:A9"].color == "#cc0000"
    assert (by_range["B1:B9"].value, by_range["B1:B9"].value2) == (2, 5.5)
    assert by_range["C1:C9"].kind == "colorscale"
    assert (by_range["C1:C9"].value, by_range["C1:C9"].value2) == ("#0000ff", "#00ff00")
    assert by_range["D1:D9"].kind == "colorscale3"
    assert by_range["D1:D9"].color == "#ffffff"
    assert by_range["E1:E9"].kind == "contains" and by_range["E1:E9"].value == "foo"
    assert by_range["F1:F9"].kind == "blank"
    assert by_range["G1:G9"].kind == "above_avg"
    assert by_range["H1:H9"].kind == "top_n" and by_range["H1:H9"].value == 3
    assert by_range["I1:I9"].kind == "bottom_pct" and by_range["I1:I9"].value == 10
    assert by_range["J1:J9"].kind == "unique"
    assert by_range["K1:K9"].kind == "==" and by_range["K1:K9"].value == "foo"


def test_xlsx_condformat_css_and_regex(tmp_path):
    pytest.importorskip("openpyxl")
    from abax.core.format.condformat import CondRule, parse_css

    wb = Workbook()
    s = wb.sheet
    s.set("A1", "7")
    s.cond_rules = [
        CondRule(range="A1:A9", kind="!=", value=7,
                 css="color: white; background: #c00; font-weight: bold"),
        CondRule(range="B1:B9", kind="regex", value="a.c"),  # no Excel counterpart
    ]
    s2 = _xlsx_roundtrip(wb, tmp_path)
    assert [r.kind for r in s2.cond_rules] == ["!="]  # regex skipped, not mangled
    # The css survives as an equivalent declaration (colours normalised to hex).
    assert parse_css(s2.cond_rules[0].css) == parse_css(s.cond_rules[0].css)


def test_xlsx_unstyled_workbook_roundtrips_clean(tmp_path):
    pytest.importorskip("openpyxl")
    wb = Workbook()
    s = wb.sheet
    s.set("A1", "5")
    s.set("B1", "=A1*2")
    s2 = _xlsx_roundtrip(wb, tmp_path)
    assert s2.get_raw(0, 1) == "=A1*2"
    assert s2.cell_formats == {} and s2.cell_styles == {} and s2.cell_borders == {}
    assert s2.col_widths == {} and s2.row_heights == {} and s2.merges == []
    assert (s2.frozen_rows, s2.frozen_cols) == (0, 0)
    assert s2.cond_rules == []


def test_xlsx_values_mode_still_carries_formatting(tmp_path):
    pytest.importorskip("openpyxl")
    from abax.engine.excel_io import load_xlsx, save_xlsx

    wb = Workbook()
    s = wb.sheet
    s.set("A1", "2")
    s.set("B1", "=A1*3")
    s.cell_formats[(0, 1)] = "fixed2"
    s.set_frozen(1, 0)
    p = tmp_path / "values.xlsx"
    save_xlsx(wb, p, values=True)
    s2 = load_xlsx(p).sheet
    assert s2.get_raw(0, 1) == "6"          # computed value, not the formula
    assert s2.cell_formats[(0, 1)] == "fixed2"
    assert (s2.frozen_rows, s2.frozen_cols) == (1, 0)


# --- embedded charts through .xlsx (line/bar/scatter map; others skip) --------


def test_xlsx_charts_roundtrip(tmp_path):
    pytest.importorskip("openpyxl")
    from abax.core.chartobj import ChartObject

    wb = Workbook()
    s = wb.sheet
    # bar source: header row + text category column
    s.set("A1", "Cat")
    s.set("B1", "Val")
    for i, (cat, val) in enumerate([("a", "3"), ("b", "5"), ("c", "2")], start=2):
        s.set(f"A{i}", cat)
        s.set(f"B{i}", val)
    # numeric block (no header) for the line and scatter charts
    for i in range(1, 6):
        s.set(f"D{i}", str(i))
        s.set(f"E{i}", str(i * i))
    data = wb.add_sheet("Data")
    for i in range(1, 5):
        data.set(f"A{i}", str(i))
        data.set(f"B{i}", str(10 * i))
    s.charts = [
        ChartObject(id="chart1", kind="bar", source="A1:B4", title="Bars",
                    anchor=(0, 6), width=400, height=300),
        ChartObject(id="chart2", kind="line", source="D1:E5", title="Squares",
                    anchor=(10, 6), width=480, height=320,
                    options={"first_col_x": True}),
        ChartObject(id="chart3", kind="scatter", source="D1:E5", title="Cloud",
                    anchor=(20, 6), width=512, height=256),
        ChartObject(id="chart4", kind="line", source="Data!A1:B4",
                    anchor=(30, 0)),
    ]
    s2 = _xlsx_roundtrip(wb, tmp_path)
    assert [(c.kind, c.source, c.title, c.anchor) for c in s2.charts] == [
        ("bar", "A1:B4", "Bars", (0, 6)),
        ("line", "D1:E5", "Squares", (10, 6)),
        ("scatter", "D1:E5", "Cloud", (20, 6)),
        ("line", "Data!A1:B4", "", (30, 0)),
    ]
    # px -> cm -> EMU -> px inverts exactly for the size
    assert [(c.width, c.height) for c in s2.charts] == [
        (400, 300), (480, 320), (512, 256), (480, 320)]
    # the x/y pairing survives as first_col_x; plain lines don't gain it
    assert s2.charts[1].options.get("first_col_x") is True
    assert s2.charts[3].options == {}
    assert [c.id for c in s2.charts] == ["chart1", "chart2", "chart3", "chart4"]


def test_xlsx_unmapped_chart_kinds_skipped(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    from abax.core.chartobj import CHART_KINDS, ChartObject
    from abax.engine.excel_io import save_xlsx

    wb = Workbook()
    s = wb.sheet
    for i in range(1, 7):
        s.set(f"A{i}", str(i))
    unmapped = [k for k in CHART_KINDS if k not in ("line", "bar", "scatter")]
    s.charts = [ChartObject(id=f"chart{n}", kind=k, source="A1:A6")
                for n, k in enumerate(unmapped, start=1)]
    s.charts.append(ChartObject(id="chart9", kind="bar", source="A1:A6",
                                title="Kept"))
    s2 = _xlsx_roundtrip(wb, tmp_path)
    # only the mappable chart comes back; the cell data always lands
    assert [(c.kind, c.source, c.title) for c in s2.charts] == \
        [("bar", "A1:A6", "Kept")]
    assert s2.get_raw(5, 0) == "6"
    # and the file itself holds exactly one native chart — nothing errored
    p = tmp_path / "native-count.xlsx"
    save_xlsx(wb, p)
    assert len(openpyxl.load_workbook(p).active._charts) == 1


def test_xlsx_foreign_chart_imports(tmp_path):
    openpyxl = pytest.importorskip("openpyxl")
    from openpyxl.chart import BarChart, PieChart, Reference

    from abax.core.chartobj import chart_data
    from abax.engine.excel_io import load_xlsx

    wb_x = openpyxl.Workbook()
    ws = wb_x.active
    for row in [["Month", "Sales"], ["Jan", 10], ["Feb", 20], ["Mar", 15]]:
        ws.append(row)
    chart = BarChart()
    chart.title = "Sales by month"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_col=2, max_row=4),
                   titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_col=1, max_row=4))
    ws.add_chart(chart, "E3")
    pie = PieChart()  # no abax counterpart -> ignored on import
    pie.add_data(Reference(ws, min_col=2, min_row=2, max_row=4))
    ws.add_chart(pie, "E20")
    p = tmp_path / "foreign.xlsx"
    wb_x.save(p)

    wb = load_xlsx(p)
    assert len(wb.sheet.charts) == 1
    ch = wb.sheet.charts[0]
    assert ch.id == "chart1"
    # header and label column fold back into one bounding source range
    assert (ch.kind, ch.source, ch.title, ch.anchor) == \
        ("bar", "A1:B4", "Sales by month", (2, 4))
    assert ch.width > 0 and ch.height > 0
    # the imported chart shapes through abax's own pipeline
    d = chart_data(wb, wb.sheet.name, ch)
    assert d["categories"] == ["Jan", "Feb", "Mar"]
    assert d["values"] == [10.0, 20.0, 15.0]


def test_xlsx_chart_sizes_convert_both_ways(tmp_path):
    pytest.importorskip("openpyxl")
    from openpyxl.utils.units import cm_to_EMU

    from abax.core.chartobj import ChartObject
    from abax.engine.excel_io import _emu_to_px, _px_to_cm

    # unit trip px -> cm -> EMU -> px is the identity (an EMU is ~0.0001 px)
    for px in (37, 100, 217, 320, 333, 480, 512, 799, 1024):
        assert _emu_to_px(cm_to_EMU(_px_to_cm(px))) == px
    # and through a real file, odd sizes land within a pixel
    wb = Workbook()
    s = wb.sheet
    for i in range(1, 5):
        s.set(f"A{i}", str(i))
    s.charts = [ChartObject(id="chart1", kind="line", source="A1:A4",
                            width=333, height=217)]
    s2 = _xlsx_roundtrip(wb, tmp_path)
    assert abs(s2.charts[0].width - 333) <= 1
    assert abs(s2.charts[0].height - 217) <= 1
